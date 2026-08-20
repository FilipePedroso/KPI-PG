#!/usr/bin/env python3
"""
Gera docs/data.json e public/data.json a partir de dois arquivos .xlsx em data/:

  - Dados.xlsx      -> aba f_venda_total (dados de venda)
  - Estrutura.xlsx  -> abas d_comercial (estrutura) e d_metas (metas)

Relacionamento: RV + ds_uf (comercial) <-> cd_vendedor + ds_uf (vendas)
                                       <-> RV + Uf (metas)
"""
import json
import sys
import os
import subprocess
from urllib.request import urlopen, Request
from urllib.error import HTTPError
from datetime import datetime, timezone, timedelta
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "data")
OUT_PATHS = [os.path.join(ROOT, "docs", "data.json"),
             os.path.join(ROOT, "public", "data.json")]

# EANs válidos para o card "Positivação Always Noturno"
ALWAYS_EANS = {
    "7506309805498", "7500435214650", "7500435214667", "7500435248334",
    "7506339326055", "7506339394603", "7500435190640", "7506339326031",
    "7500435265263", "7500435233446", "7506339325263", "7500435190657",
    "7506339394535", "7506339325249",
}

# Bitmask dos indicadores usados na visão "Clientes"
# (cards de Positivação, Ranking e Faturamento)
CLIENT_METRICS = ["tot", "ali", "far", "hfs", "rfar", "alw", "pmp", "fec", "fsp",
                  "ecp", "ch2"]
CB = {m: 1 << i for i, m in enumerate(CLIENT_METRICS)}
# Métricas com valor financeiro por cliente (ordem do array de valores)
CLIENT_VALUES = ["tot", "ali", "far", "fec", "fsp"]
FARMA_CHANNELS = {"DRUG/PHARMACY", "PERFUMERIES"}



def asset_base_url():
    """Retorna a URL base do projeto para resolver assets relativos."""
    return os.environ.get("LOVABLE_PROJECT_URL",
                          "https://id-preview--fcb0edde-deb7-44dc-ab51-3bee86aa48cf.lovable.app")


def read_asset(name):
    p = os.path.join(DATA_DIR, name + ".asset.json")
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def write_asset(name, payload):
    p = os.path.join(DATA_DIR, name + ".asset.json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  -> asset pointer atualizado: {p}")


def upload_asset(name, source):
    """Faz upload do arquivo local e retorna o asset pointer atualizado."""
    try:
        result = subprocess.run(
            ["lovable-assets", "create", "--file", source],
            capture_output=True, text=True, check=True,
        )
        payload = json.loads(result.stdout)
        write_asset(name, payload)
        return payload
    except Exception as e:
        print(f"  AVISO: não foi possível fazer upload de {name}: {e}")
        return None


def download_asset(name, target, asset):
    """Baixa o asset do CDN para o path local."""
    url = asset.get("url", "")
    if not url:
        return False
    if url.startswith("/"):
        url = asset_base_url().rstrip("/") + url
    try:
        print(f"  -> baixando {name} de {url} ...")
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 (Lovable build)"})
        with urlopen(req) as resp:
            with open(target, "wb") as f:
                f.write(resp.read())
        return True
    except HTTPError as e:
        print(f"  ERRO ao baixar {name}: {e.code} {e.reason}")
        return False
    except Exception as e:
        print(f"  ERRO ao baixar {name}: {e}")
        return False



def ensure_asset(name):
    """
    Garante que o arquivo .xlsx esteja disponível em data/{name}.
    - Se existe localmente e mudou de tamanho, faz upload e atualiza pointer.
    - Se não existe localmente, baixa do CDN usando o pointer.
    Retorna o path local ou None.
    """
    source = os.path.join(DATA_DIR, name)
    asset = read_asset(name)

    if os.path.exists(source):
        local_size = os.path.getsize(source)
        asset_size = asset.get("size") if asset else None
        if os.environ.get("CI"):
            # No GitHub Actions usamos o .xlsx do próprio repositório.
            print(f"{name}: usando arquivo do repositório ({local_size:,} bytes).")
            return source
        if asset_size is None or local_size != asset_size:
            print(f"{name} local ({local_size:,} bytes) difere do asset ({asset_size or '?'}). Fazendo upload...")
            new_asset = upload_asset(name, source)
            if new_asset:
                return source
        return source

    if asset:
        if download_asset(name, source, asset):
            return source
        print(f"ERRO: {name} não encontrado localmente e não pôde ser baixado do CDN.")
        return None

    print(f"ERRO: {name} não encontrado localmente e não existe asset pointer.")
    return None






def s(v):
    return "" if v is None else str(v).strip()


def n(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def rv_key(v):
    """Normaliza código de vendedor (448 / '448' / 448.0 -> '448')."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    t = str(v).strip()
    if t.endswith(".0"):
        t = t[:-2]
    return t


def header_map(ws):
    rows = ws.iter_rows(values_only=True)
    header = [s(h) for h in next(rows)]
    idx = {}
    for i, h in enumerate(header):
        if h and h not in idx:
            idx[h] = i
        low = h.strip().lower()
        idx.setdefault("~" + low, i)
    return rows, header, idx


def col(idx, *names):
    for nm in names:
        if nm in idx:
            return idx[nm]
        k = "~" + nm.strip().lower()
        if k in idx:
            return idx[k]
    return None


# Arquivos parquet das tabelas fato (substituem as abas do Dados.xlsx)
PARQUET_FILES = {
    "f_venda_total": "Dados_f_venda_total.parquet",
    "f_ec_oniz": "Dados_f_ec_oniz.parquet",
}


def parquet_rows(path):
    """Lê um parquet e devolve (rows, header, idx) no mesmo formato de header_map."""
    import pyarrow.parquet as pq
    table = pq.read_table(path)
    header = [s(h) for h in table.column_names]
    idx = {}
    for i, h in enumerate(header):
        if h and h not in idx:
            idx[h] = i
        idx.setdefault("~" + h.strip().lower(), i)
    cols = [c.to_pylist() for c in table.columns]
    rows = (tuple(c[i] for c in cols) for i in range(table.num_rows))
    return rows, header, idx


def fact_source(sheet, wbd=None):
    """Prefere o parquet em data/; cai para a aba do Dados.xlsx se não existir."""
    p = os.path.join(DATA_DIR, PARQUET_FILES[sheet])
    if os.path.exists(p):
        print(f"[build_data] {sheet}: usando {os.path.basename(p)}", file=sys.stderr)
        return parquet_rows(p)
    if wbd is not None and sheet in wbd.sheetnames:
        return header_map(wbd[sheet])
    return None


SC_FILE = "Dados_SC.xlsx"


def read_sc():
    """
    Lê data/Dados_SC.xlsx (indicadores já tratados de SC) e devolve
    {rv|uf: {hfs, far, alw, pmp, ch2, pp}} somando a coluna POSITIVAÇÃO.
    """
    path = os.path.join(DATA_DIR, SC_FILE)
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    acc = {}

    def bucket(rv, uf):
        rv, uf = rv_key(rv), s(uf)
        if not uf:
            return None
        # linhas sem vendedor identificado (ex.: "Outros") vão para o código 0
        if not rv or not rv.isdigit():
            rv = "0"
        return acc.setdefault(rv + "|" + uf, {"hfs": 0.0, "far": 0.0,
                                              "alw": 0.0, "pmp": 0.0,
                                              "ch2": 0.0, "pp": 0.0})


    def sheet(name):
        return header_map(wb[name]) if name in wb.sheetnames else (None, None, None)

    def pos_col(header, last=False):
        """Índice da coluna POSITIVAÇÃO (a última quando last=True)."""
        hits = [i for i, h in enumerate(header)
                if s(h).strip().lower().startswith("positiva")]
        if not hits:
            return None
        return hits[-1] if last else hits[0]

    # Pos Relação -> HFS / FARMA
    rows, hdr, idx = sheet("Pos Relação")
    if rows:
        c_rv, c_uf = col(idx, "RV", "cd_vendedor"), col(idx, "ds_uf")
        c_tp, c_pos = col(idx, "TIPO"), pos_col(hdr)
        for r in rows:
            if not r:
                continue
            b = bucket(r[c_rv], r[c_uf])
            if not b:
                continue
            val = n(r[c_pos]) if c_pos is not None else 0
            tp = s(r[c_tp]).upper()
            if tp == "HFS":
                b["hfs"] += val
            elif tp == "FARMA":
                b["far"] += val

    # Marcas Relação -> ALWAYS NOTURNO / PAMPERS
    rows, hdr, idx = sheet("Marcas Relação")
    if rows:
        c_rv, c_uf = col(idx, "RV", "cd_vendedor"), col(idx, "ds_uf")
        c_tp, c_pos = col(idx, "TIPO"), pos_col(hdr)
        for r in rows:
            if not r:
                continue
            b = bucket(r[c_rv], r[c_uf])
            if not b:
                continue
            val = n(r[c_pos]) if c_pos is not None else 0
            tp = s(r[c_tp]).upper()
            if tp == "ALWAYS NOTURNO":
                b["alw"] += val
            elif tp == "PAMPERS":
                b["pmp"] += val

    # Escolha Certa -> chaves >= 2
    rows, hdr, idx = sheet("Escolha Certa")
    if rows:
        c_rv, c_uf = col(idx, "cd_vendedor", "RV"), col(idx, "ds_uf")
        c_pos = pos_col(hdr)
        for r in rows:
            if not r:
                continue
            b = bucket(r[c_rv], r[c_uf])
            if b and c_pos is not None:
                b["ch2"] += n(r[c_pos])

    # Platinum Points -> segunda coluna POSITIVAÇÃO
    rows, hdr, idx = sheet("Platinum Points")
    if rows:
        c_rv, c_uf = col(idx, "Rv", "RV", "cd_vendedor"), col(idx, "ds_uf")
        c_pos = pos_col(hdr, last=True)
        for r in rows:
            if not r:
                continue
            b = bucket(r[c_rv], r[c_uf])
            if b and c_pos is not None:
                b["pp"] += n(r[c_pos])

    return {k: {m: round(v) for m, v in b.items()} for k, b in acc.items()}


def strip_faixa(v):
    """'faixa3' -> '3' (remove o prefixo textual da coluna FAIXA de SC)."""
    t = s(v)
    if t.lower().startswith("faixa"):
        t = t[len("faixa"):]
    return t.strip(" -:_")


def read_sc_clientes():
    """
    Lê data/Dados_SC.xlsx e devolve, por rv|uf, os CNPJs positivados de cada
    indicador de ranking, além do nome (razão social) e, para Chaves >= 2,
    a faixa (chave) atingida por CNPJ em cada vendedor (um mesmo CNPJ pode
    aparecer sob mais de um vendedor ao longo do tempo).
      -> ({rv|uf: {hfs|rfar|alw|pmp|ch2: set(cnpj)}}, {cnpj: nome},
          {rv|uf: {cnpj: chave}})
    """
    path = os.path.join(DATA_DIR, SC_FILE)
    if not os.path.exists(path):
        return {}, {}, {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    acc, nomes, faixas = {}, {}, {}

    def key(rv, uf):
        rv, uf = rv_key(rv), s(uf)
        if not uf:
            return None
        if not rv or not rv.isdigit():
            rv = "0"
        return rv + "|" + uf

    def scan(sheet, tipos=None, fixed_metric=None, extra_col=None, extra_dict=None):
        if sheet not in wb.sheetnames:
            return
        rows, hdr, idx = header_map(wb[sheet])
        c_rv = col(idx, "RV", "Rv", "cd_vendedor")
        c_uf = col(idx, "ds_uf")
        c_tp = col(idx, "TIPO")
        c_cnpj = col(idx, "CNPJ")
        c_nome = col(idx, "RAZÃO SOCIAL")
        c_extra = col(idx, extra_col) if extra_col else None
        hits = [i for i, h in enumerate(hdr)
                if s(h).strip().lower().startswith("positiva")]
        c_pos = hits[0] if hits else None
        for r in rows:
            if not r:
                continue
            k = key(r[c_rv], r[c_uf])
            cnpj = rv_key(r[c_cnpj]) if c_cnpj is not None else ""
            if not k or not cnpj:
                continue
            if c_pos is not None and n(r[c_pos]) <= 0:
                continue
            metric = fixed_metric if fixed_metric else \
                tipos.get(s(r[c_tp]).upper() if c_tp is not None else "")
            if not metric:
                continue
            acc.setdefault(k, {}).setdefault(metric, set()).add(cnpj)
            if c_nome is not None and cnpj not in nomes:
                nomes[cnpj] = s(r[c_nome])
            if c_extra is not None and extra_dict is not None:
                kdict = extra_dict.setdefault(k, {})
                if cnpj not in kdict:
                    kdict[cnpj] = strip_faixa(r[c_extra])

    scan("Pos Relação", tipos={"HFS": "hfs", "FARMA": "rfar"})
    scan("Marcas Relação", tipos={"ALWAYS NOTURNO": "alw", "PAMPERS": "pmp"})
    scan("Escolha Certa", fixed_metric="ch2", extra_col="FAIXA", extra_dict=faixas)
    return acc, nomes, faixas






def read_competencia(wbe):
    """
    Lê a aba 'data' de Estrutura.xlsx: célula única com a data de competência
    (mês/ano usado como referência para as médias diárias dos cards).
    """
    if "data" not in wbe.sheetnames:
        return None
    ws = wbe["data"]
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row or ():
            if isinstance(cell, datetime):
                return cell
            if cell:
                try:
                    return datetime.fromisoformat(s(cell))
                except ValueError:
                    continue
    return None


def build(dados_path, estrutura_path):
    # ---------- Estrutura: d_comercial ----------
    wbe = openpyxl.load_workbook(estrutura_path, read_only=True, data_only=True)
    competencia = read_competencia(wbe)
    ws = wbe["d_comercial"]
    rows, header, idx = header_map(ws)
    i_rv = col(idx, "RV")
    i_uf = col(idx, "ds_uf")
    i_rvn = col(idx, "CONCATENAÇÃO RV + NOME")
    i_svn = col(idx, "CONCATENAÇÃO SV + NOME")
    i_cvn = col(idx, "CONCATENAÇÃO CV + NOME")
    comercial, seen = [], set()
    for r in rows:
        if not r:
            continue
        rv = rv_key(r[i_rv])
        uf = s(r[i_uf])
        if not rv or not uf:
            continue
        k = rv + "|" + uf
        if k in seen:
            continue
        seen.add(k)
        comercial.append({
            "rv": rv, "uf": uf,
            "rvName": s(r[i_rvn]),
            "sv": s(r[i_svn]),
            "cv": s(r[i_cvn]),
        })

    # ---------- Estrutura: d_metas ----------
    ws = wbe["d_metas"]
    rows, header, idx = header_map(ws)
    m_rv = col(idx, "RV")
    m_uf = col(idx, "Uf", "ds_uf")
    m_tot = col(idx, "Meta Financeira Total")
    m_ec = col(idx, "Meta Financeira Escolha Certa")
    m_sp = col(idx, "Meta Financeira Store Platform")
    m_ali = col(idx, "Meta Financeira Alimentar")
    m_far = col(idx, "Meta Financeira Farma")
    p_tot = col(idx, "Objetivo Positivação Total")
    p_ali = col(idx, "Objetivo Positivação Alimentar")
    p_far = col(idx, "Objetivo Positivação Farma")
    r_hfs = col(idx, "OBJ PRODUTIVIDADE HFS")
    r_far = col(idx, "OBJ PRODUTIVIDADE FARMA", "OBJ PRODUTIVIDADE FARMA ")
    r_alw = col(idx, "Objetivo Marca 1")
    r_pmp = col(idx, "Objetivo Marca 2")
    e_ch2 = col(idx, "OBJ ESCOLHA CERTA_NOVO")
    e_pp = col(idx, "Objetivo Platinum Points")
    magg = {}
    for r in rows:
        if not r:
            continue
        rv = rv_key(r[m_rv])
        uf = s(r[m_uf])
        if not rv or not uf:
            continue
        k = rv + "|" + uf
        b = magg.setdefault(k, {"rv": rv, "uf": uf, "total": 0.0, "ec": 0.0,
                                "sp": 0.0, "ali": 0.0, "far": 0.0,
                                "p_total": 0.0, "p_ali": 0.0, "p_far": 0.0,
                                "r_hfs": 0.0, "r_far": 0.0,
                                "r_alw": 0.0, "r_pmp": 0.0,
                                "e_ch2": 0.0, "e_pp": 0.0})
        b["total"] += n(r[m_tot]) if m_tot is not None else 0.0
        b["ec"] += n(r[m_ec]) if m_ec is not None else 0.0
        b["sp"] += n(r[m_sp]) if m_sp is not None else 0.0
        b["ali"] += n(r[m_ali]) if m_ali is not None else 0.0
        b["far"] += n(r[m_far]) if m_far is not None else 0.0
        b["p_total"] += n(r[p_tot]) if p_tot is not None else 0.0
        b["p_ali"] += n(r[p_ali]) if p_ali is not None else 0.0
        b["p_far"] += n(r[p_far]) if p_far is not None else 0.0
        b["r_hfs"] += n(r[r_hfs]) if r_hfs is not None else 0.0
        b["r_far"] += n(r[r_far]) if r_far is not None else 0.0
        b["r_alw"] += n(r[r_alw]) if r_alw is not None else 0.0
        b["r_pmp"] += n(r[r_pmp]) if r_pmp is not None else 0.0
        b["e_ch2"] += n(r[e_ch2]) if e_ch2 is not None else 0.0
        b["e_pp"] += n(r[e_pp]) if e_pp is not None else 0.0
    metas = list(magg.values())

    # ---------- Estrutura: d_clientes_braveo (potencial + base de clientes) ----------
    # Contagem distinta de CNPJ por rv|uf, aplicando as mesmas regras de
    # negócio (Canal Ranking / Plataforma) de cada card de ranking.
    potencial = {}
    clientes_base = {}   # {rv|uf: {cnpj: [nome, elig_bitmask]}}
    if "d_clientes_braveo" in wbe.sheetnames:
        ws = wbe["d_clientes_braveo"]
        rows, header, idx = header_map(ws)
        c_rv = col(idx, "cd_vendedor")
        c_uf = col(idx, "ds_uf")
        c_cnpj = col(idx, "nr_cnpj_cpf")
        c_nome = col(idx, "nm_pessoa")
        c_chan = col(idx, "Store Channel")
        c_can = col(idx, "Canal Ranking")
        c_plat = col(idx, "Plataforma")
        psets = {}
        for r in rows:
            if not r:
                continue
            rv = rv_key(r[c_rv])
            uf = s(r[c_uf])
            cnpj = s(r[c_cnpj]) if c_cnpj is not None else ""
            if not rv or not uf or not cnpj:
                continue
            canal = s(r[c_can]) if c_can is not None else ""
            plat = s(r[c_plat]) if c_plat is not None else ""
            chan = s(r[c_chan]).upper() if c_chan is not None else ""
            k = rv + "|" + uf
            b = psets.setdefault(k, {"hfs": set(), "far": set(),
                                     "alw": set(), "pmp": set(), "ecp": set()})
            plat_ok = plat in ("Escolha Certa", "Store Platform")
            if canal == "HFS" and plat_ok:
                b["hfs"].add(cnpj)
            if canal in ("Farma Indep", "Farma Rede") and plat_ok:
                b["far"].add(cnpj)
            if canal in ("HFS", "Farma Indep") and plat == "Escolha Certa":
                b["alw"].add(cnpj)
            if canal in ("HFS", "Farma Indep") and plat_ok:
                b["pmp"].add(cnpj)
            if plat == "Escolha Certa":
                b["ecp"].add(cnpj)

            # base de clientes (para a visão "Clientes" dos cards)
            elig = CB["tot"]
            elig |= CB["far"] if chan in FARMA_CHANNELS else CB["ali"]
            if canal == "HFS" and plat_ok:
                elig |= CB["hfs"]
            if canal in ("Farma Indep", "Farma Rede") and plat_ok:
                elig |= CB["rfar"]
            if canal in ("HFS", "Farma Indep") and plat == "Escolha Certa":
                elig |= CB["alw"]
            if canal in ("HFS", "Farma Indep") and plat_ok:
                elig |= CB["pmp"]
            if plat == "Escolha Certa":
                elig |= CB["fec"] | CB["ecp"] | CB["ch2"]
            elif plat == "Store Platform":
                elig |= CB["fsp"]
            cb = clientes_base.setdefault(k, {})
            cur = cb.get(cnpj)
            if cur is None:
                cb[cnpj] = [s(r[c_nome]) if c_nome is not None else "", elig]
            else:
                cur[1] |= elig
        for k, b in psets.items():
            rv, uf = k.split("|", 1)
            potencial[k] = {"rv": rv, "uf": uf,
                            "hfs": len(b["hfs"]), "far": len(b["far"]),
                            "alw": len(b["alw"]), "pmp": len(b["pmp"]),
                            "ecp": len(b["ecp"])}



    # ---------- Dados: f_venda_total ----------
    wbd = None
    if dados_path:
        wbd = openpyxl.load_workbook(dados_path, read_only=True, data_only=True)
    src = fact_source("f_venda_total", wbd)
    if not src:
        print("ERRO: fonte f_venda_total não encontrada (parquet ou xlsx).")
        sys.exit(1)
    rows, header, idx = src
    print(f"[build_data] f_venda_total headers: {header}", file=sys.stderr)

    v_rv = col(idx, "cd_vendedor")
    v_uf = col(idx, "ds_uf")
    v_val = col(idx, "vl_financeiro")
    v_plat = col(idx, "Plataforma")
    v_can = col(idx, "Canal")
    v_fat = col(idx, "vl_faturamento")
    v_cnpj = col(idx, "nr_cnpj_cpf")
    v_pessoa = col(idx, "nm_pessoa")
    v_ger = col(idx, "cd_gerente")
    v_sup = col(idx, "cd_vendedor_superior")
    v_crank = col(idx, "Canal Ranking")
    v_ean = col(idx, "ds_ean")
    v_grupo = col(idx, "nm_grupo")
    v_prod = col(idx, "nm_produto")
    v_vol = col(idx, "vl_volume_cx")

    vagg = {}
    cnpj_sums = {}
    fato_nomes = {}   # {cnpj: nm_pessoa} - fallback de nome p/ clientes fora da base
    rank_sums = {}
    total_rows = fat_rows = 0

    for r in rows:
        if not r:
            continue
        rv = rv_key(r[v_rv])
        uf = s(r[v_uf])
        if not rv or not uf:
            continue
        val = n(r[v_val])
        plat = s(r[v_plat]) if v_plat is not None else ""
        canal = s(r[v_can]).lower() if v_can is not None else ""
        is_far = canal == "farma"
        raw = r[v_fat] if v_fat is not None else None
        if isinstance(raw, (int, float)):
            is_fat = raw != 0
        else:
            sv = s(raw).lower()
            is_fat = sv not in ("", "0", "0.0", "nao", "não", "no", "false")
        total_rows += 1
        if is_fat:
            fat_rows += 1
        k = rv + "|" + uf
        b = vagg.setdefault(k, {
            "rv": rv, "uf": uf,
            "v": 0.0, "ec": 0.0, "sp": 0.0, "ali": 0.0, "far": 0.0,
            "vf": 0.0, "vf_ec": 0.0, "vf_sp": 0.0, "vf_ali": 0.0, "vf_far": 0.0,
            "p": 0, "p_ali": 0, "p_far": 0,
            "pf": 0, "pf_ali": 0, "pf_far": 0,
        })
        if v_sup is not None and not b.get("cdSv"):
            b["cdSv"] = rv_key(r[v_sup])
        if v_ger is not None and not b.get("cdCv"):
            b["cdCv"] = rv_key(r[v_ger])
        b["v"] += val
        if plat == "Escolha Certa":
            b["ec"] += val
        elif plat == "Store Platform":
            b["sp"] += val
        if is_far:
            b["far"] += val
        else:
            b["ali"] += val
        if is_fat:
            b["vf"] += val
            if plat == "Escolha Certa":
                b["vf_ec"] += val
            elif plat == "Store Platform":
                b["vf_sp"] += val
            if is_far:
                b["vf_far"] += val
            else:
                b["vf_ali"] += val
        cnpj = s(r[v_cnpj]) if v_cnpj is not None else ""
        if cnpj:
            if v_pessoa is not None and cnpj not in fato_nomes:
                nm = s(r[v_pessoa])
                if nm:
                    fato_nomes[cnpj] = nm
            cs = cnpj_sums.setdefault(k, {}).setdefault(
                cnpj, {"t": 0.0, "a": 0.0, "f": 0.0, "tf": 0.0, "af": 0.0,
                       "ff": 0.0, "ec": 0.0, "sp": 0.0})
            cs["t"] += val
            cs["f" if is_far else "a"] += val
            if plat == "Escolha Certa":
                cs["ec"] += val
            elif plat == "Store Platform":
                cs["sp"] += val
            if is_fat:
                cs["tf"] += val
                cs["ff" if is_far else "af"] += val

            # ----- Ranking (Canal Ranking / Plataforma / EAN / grupo) -----
            crank = s(r[v_crank]) if v_crank is not None else ""
            plat_ok = plat in ("Escolha Certa", "Store Platform")
            ean = rv_key(r[v_ean]) if v_ean is not None else ""
            grupo = s(r[v_grupo]).upper() if v_grupo is not None else ""
            prod = s(r[v_prod]).upper() if v_prod is not None else ""
            vol = n(r[v_vol]) if v_vol is not None else 0.0
            rb = rank_sums.setdefault(k, {})

            def add_rank(metric):
                cc = rb.setdefault(metric, {}).setdefault(
                    cnpj, {"t": 0.0, "f": 0.0, "vt": 0.0, "vf": 0.0})
                cc["t"] += val
                cc["vt"] += vol
                if is_fat:
                    cc["f"] += val
                    cc["vf"] += vol

            if crank == "HFS" and plat_ok:
                add_rank("hfs")
            if crank in ("Farma Indep", "Farma Rede") and plat_ok:
                add_rank("far")
            if crank in ("HFS", "Farma Indep") and plat == "Escolha Certa" \
                    and ean in ALWAYS_EANS:
                add_rank("alw")
            if crank in ("HFS", "Farma Indep") and plat_ok \
                    and ("PAMPERS" in grupo or "FRALDA" in grupo) \
                    and "TOALHAS" not in prod:
                add_rank("pmp")
            if plat == "Escolha Certa":
                add_rank("ecp")


    pos_total_all = 0
    for k, cm in cnpj_sums.items():
        b = vagg.get(k)
        if not b:
            continue
        for cs in cm.values():
            if cs["t"] > 0:
                b["p"] += 1
                pos_total_all += 1
            if cs["a"] > 0:
                b["p_ali"] += 1
            if cs["f"] > 0:
                b["p_far"] += 1
            if cs["tf"] > 0:
                b["pf"] += 1
            if cs["af"] > 0:
                b["pf_ali"] += 1
            if cs["ff"] > 0:
                b["pf_far"] += 1

    for k, mm in rank_sums.items():
        b = vagg.get(k)
        if not b:
            continue
        for metric, cm in mm.items():
            minv = 10.0 if metric == "pmp" else 0.0
            tot = sum(1 for cs in cm.values() if cs["t"] > 0 and cs["vt"] >= minv)
            fat = sum(1 for cs in cm.values() if cs["f"] > 0 and cs["vf"] >= minv)
            b["r_" + metric] = tot
            b["rf_" + metric] = fat

    vendas = list(vagg.values())
    print(f"[build_data] linhas: {total_rows}, faturadas: {fat_rows}, "
          f"CNPJs positivados: {pos_total_all}", file=sys.stderr)

    # ---------- Dados: f_ec_oniz (Escolha Certa) ----------
    # chaves >= 2: contagem distinta de nr_doc com nr_chave >= 2
    # platinum points: contagem distinta de (nr_doc, ds_combo_sku_lista_ativacao)
    #                  com "Platinum Point?" = "Sim"
    ec = []
    ec_ch2_cli = {}     # {rv|uf: set(cnpj)} - clientes com chaves >= 2 (nr_doc == CNPJ)
    ec_nomes = {}       # {cnpj: nm_pessoa}
    ec_chave_by_key = {}  # {rv|uf: {cnpj: maior nr_chave atingido NAQUELE vendedor}}
    src_ec = fact_source("f_ec_oniz", wbd)
    if src_ec:
        rows, header, idx = src_ec

        o_rv = col(idx, "cd_vendedor")
        o_uf = col(idx, "ds_sigla", "ds_uf")
        o_doc = col(idx, "nr_doc")
        o_ch = col(idx, "nr_chave")
        o_combo = col(idx, "ds_combo_sku_lista_ativacao")
        o_pp = col(idx, "Platinum Point?")
        o_nome = col(idx, "nm_pessoa")
        esets = {}
        for r in rows:
            if not r:
                continue
            rv = rv_key(r[o_rv])
            uf = s(r[o_uf])
            doc = rv_key(r[o_doc]) if o_doc is not None else ""
            if not rv or not uf or not doc:
                continue
            k = rv + "|" + uf
            b = esets.setdefault(k, {"ch2": set(), "pp": set()})
            if o_ch is not None:
                chv = int(n(r[o_ch]))
                kchave = ec_chave_by_key.setdefault(k, {})
                if chv > kchave.get(doc, 0):
                    kchave[doc] = chv
                if chv >= 2:
                    b["ch2"].add(doc)
                if o_nome is not None and doc not in ec_nomes:
                    nm = s(r[o_nome])
                    if nm:
                        ec_nomes[doc] = nm
            if o_pp is not None and s(r[o_pp]).lower() in ("sim", "s", "yes", "1"):
                combo = s(r[o_combo]) if o_combo is not None else ""
                b["pp"].add(doc + "|" + combo)
        for k, b in esets.items():
            rv, uf = k.split("|", 1)
            ec.append({"rv": rv, "uf": uf,
                       "ch2": len(b["ch2"]), "pp": len(b["pp"])})
            if b["ch2"]:
                ec_ch2_cli[k] = set(b["ch2"])
        print(f"[build_data] f_ec_oniz: chaves>=2 = "
              f"{sum(e['ch2'] for e in ec)}, platinum = "
              f"{sum(e['pp'] for e in ec)}", file=sys.stderr)

    # ---------- Dados_SC.xlsx (fonte alternativa para as UFs contidas nele) ----------
    sc = read_sc()
    if sc:
        sc_ufs = {k.split("|", 1)[1] for k in sc}
        vmap = {b["rv"] + "|" + b["uf"]: b for b in vendas}
        emap = {e["rv"] + "|" + e["uf"]: e for e in ec}
        keys = set(sc) | {k for k in vmap if k.split("|", 1)[1] in sc_ufs} \
                       | {k for k in emap if k.split("|", 1)[1] in sc_ufs}
        for k in keys:
            rv, uf = k.split("|", 1)
            v = sc.get(k, {})
            b = vmap.get(k)
            if b is None:
                b = {"rv": rv, "uf": uf,
                     "v": 0.0, "ec": 0.0, "sp": 0.0, "ali": 0.0, "far": 0.0,
                     "vf": 0.0, "vf_ec": 0.0, "vf_sp": 0.0, "vf_ali": 0.0,
                     "vf_far": 0.0, "p": 0, "p_ali": 0, "p_far": 0,
                     "pf": 0, "pf_ali": 0, "pf_far": 0}
                vendas.append(b)
                vmap[k] = b
            for m in ("hfs", "far", "alw", "pmp"):
                b["r_" + m] = v.get(m, 0)
                b["rf_" + m] = v.get(m, 0)
            e = emap.get(k)
            if e is None:
                e = {"rv": rv, "uf": uf, "ch2": 0, "pp": 0}
                ec.append(e)
                emap[k] = e
            e["ch2"] = v.get("ch2", 0)
            e["pp"] = v.get("pp", 0)
        print(f"[build_data] Dados_SC: UFs {sorted(sc_ufs)}, "
              f"{len(sc)} combinações rv|uf sobrescritas", file=sys.stderr)

    # ---------- Base de clientes por indicador (visão "Clientes") ----------
    # pos[k][metric] = set(cnpj positivados)
    pos = {}
    for k, cm in cnpj_sums.items():
        b = pos.setdefault(k, {})
        for cnpj, cs in cm.items():
            if cs["t"] > 0:
                b.setdefault("tot", set()).add(cnpj)
            if cs["a"] > 0:
                b.setdefault("ali", set()).add(cnpj)
            if cs["f"] > 0:
                b.setdefault("far", set()).add(cnpj)
            if cs["ec"] > 0:
                b.setdefault("fec", set()).add(cnpj)
            if cs["sp"] > 0:
                b.setdefault("fsp", set()).add(cnpj)
    rank_alias = {"hfs": "hfs", "far": "rfar", "alw": "alw", "pmp": "pmp",
                  "ecp": "ecp"}
    for k, mm in rank_sums.items():
        b = pos.setdefault(k, {})
        for metric, cm in mm.items():
            alias = rank_alias.get(metric)
            if not alias:
                continue
            minv = 10.0 if metric == "pmp" else 0.0
            for cnpj, cs in cm.items():
                if cs["t"] > 0 and cs["vt"] >= minv:
                    b.setdefault(alias, set()).add(cnpj)
    for k, cnpjs in ec_ch2_cli.items():
        pos.setdefault(k, {})["ch2"] = set(cnpjs)

    # SC: os indicadores de ranking vêm do Dados_SC.xlsx (CNPJ a CNPJ)
    sc_cli, sc_nomes, sc_faixa = read_sc_clientes()
    if sc:
        for k in list(pos):
            if k.split("|", 1)[1] in sc_ufs:
                for alias in ("hfs", "rfar", "alw", "pmp", "ch2"):
                    pos[k].pop(alias, None)
        for k, mm in sc_cli.items():
            b = pos.setdefault(k, {})
            for alias, cnpjs in mm.items():
                b[alias] = set(cnpjs)

    clientes = {}
    for k in set(clientes_base) | set(pos):
        base = dict(clientes_base.get(k, {}))
        pk = pos.get(k, {})
        for alias, cnpjs in pk.items():
            for cnpj in cnpjs:
                if cnpj not in base:
                    # positivado fora da base de clientes -> entra marcado
                    base[cnpj] = [sc_nomes.get(cnpj) or fato_nomes.get(cnpj)
                                  or ec_nomes.get(cnpj, ""), 0]
        out = []
        vsrc = cnpj_sums.get(k, {})
        rsrc = rank_sums.get(k, {}).get("pmp", {})
        for cnpj, (nome, elig) in base.items():
            if not nome:
                nome = fato_nomes.get(cnpj) or ec_nomes.get(cnpj, "")
            posmask = 0
            for alias, bit in CB.items():
                if cnpj in pk.get(alias, ()):
                    posmask |= bit
            if not (elig | posmask):
                continue
            cs = vsrc.get(cnpj)
            vals = [round(cs["t"], 2), round(cs["a"], 2), round(cs["f"], 2),
                    round(cs["ec"], 2), round(cs["sp"], 2)] if cs else [0, 0, 0, 0, 0]
            vol = rsrc.get(cnpj, {}).get("vt", 0.0)
            raw_faixa = sc_faixa.get(k, {}).get(cnpj)
            if raw_faixa:
                try:
                    chave = int(raw_faixa)
                except ValueError:
                    chave = 0
            else:
                chave = ec_chave_by_key.get(k, {}).get(cnpj, 0)
            out.append([cnpj, nome, elig, posmask, vals, round(vol, 1), chave])
        if out:
            out.sort(key=lambda x: x[1])
            clientes[k] = out
    print(f"[build_data] clientes: {sum(len(v) for v in clientes.values())} "
          f"linhas em {len(clientes)} combinações rv|uf", file=sys.stderr)

    br = timezone(timedelta(hours=-3))

    return {
        "generated_at": datetime.now(br).strftime("%Y-%m-%dT%H:%M:%S-03:00"),
        "competencia": competencia.strftime("%Y-%m-01T00:00:00-03:00") if competencia else None,
        "source_file": os.path.basename(dados_path) if dados_path else "parquet",
        "comercial": comercial,
        "vendas": vendas,
        "metas": metas,
        "potencial": list(potencial.values()),
        "ec": ec,
        "_clientes": {"metrics": CLIENT_METRICS, "values": CLIENT_VALUES,
                      "clientes": clientes},
    }





def find(name):
    return ensure_asset(name)



def build_orfaos(data):
    """Combinações rv|uf presentes na base mas ausentes da d_comercial."""
    est = {c["rv"] + "|" + c["uf"] for c in data["comercial"]}
    itens = []
    for x in data["vendas"]:
        k = x["rv"] + "|" + x["uf"]
        if k in est:
            continue
        itens.append({
            "chave": k, "rv": x["rv"], "uf": x["uf"], "origem": "vendas",
            "vl_financeiro": round(x.get("v", 0.0), 2),
            "vl_faturado": round(x.get("vf", 0.0), 2),
            "positivados": x.get("p", 0),
            "cdSv": x.get("cdSv", ""), "cdCv": x.get("cdCv", ""),
        })
    vkeys = {x["rv"] + "|" + x["uf"] for x in data["vendas"]}
    for m in data["metas"]:
        k = m["rv"] + "|" + m["uf"]
        if k in est or k in vkeys:
            continue
        itens.append({
            "chave": k, "rv": m["rv"], "uf": m["uf"], "origem": "metas",
            "vl_financeiro": 0.0, "vl_faturado": 0.0, "positivados": 0,
            "cdSv": "", "cdCv": "",
            "meta_financeira": round(m.get("total", 0.0), 2),
            "meta_positivacao": round(m.get("p_total", 0.0), 2),
        })
    itens.sort(key=lambda i: -i.get("vl_financeiro", 0.0))
    return itens


def code_of(label):
    """Extrai o código de rótulos como '11 - NOME' ou '211-NOME'."""
    t = s(label)
    return rv_key(t.split("-")[0]) if "-" in t else rv_key(t)


def completar_estrutura(data, itens):
    """Adiciona à d_comercial as combinações órfãs, reaproveitando nomes
    já existentes de supervisor/gerente quando o código for conhecido."""
    known = {}
    for c in data["comercial"]:
        for field in ("sv", "cv"):
            lbl = c.get(field) or ""
            cd = code_of(lbl)
            if cd and "-" in lbl and lbl.split("-", 1)[1].strip() not in ("", "-"):
                known.setdefault(cd, lbl)

    def label(cd):
        if not cd:
            return "-"
        return known.get(cd, f"{cd} - -")

    for i in itens:
        data["comercial"].append({
            "rv": i["rv"], "uf": i["uf"],
            "rvName": f"{i['rv']} - -",
            "sv": label(i.get("cdSv")),
            "cv": label(i.get("cdCv")),
        })


def csv_escape(v):
    t = "" if v is None else str(v)
    return '"' + t.replace('"', '""') + '"' if (";" in t or '"' in t or "\n" in t) else t


def write_csv(name, rows, header):
    lines = [";".join(header)]
    for r in rows:
        lines.append(";".join(csv_escape(c) for c in r))
    payload = "\ufeff" + "\r\n".join(lines) + "\r\n"
    for base in ("docs", "public"):
        p = os.path.join(ROOT, base, name)
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            f.write(payload)
        print(f"  -> {p} ({len(payload):,} bytes)")


def write_all(name, obj, indent=None):
    payload = json.dumps(obj, ensure_ascii=False,
                         separators=(",", ":") if indent is None else None,
                         indent=indent)
    for base in ("docs", "public"):
        p = os.path.join(ROOT, base, name)
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            f.write(payload)
        print(f"  -> {p} ({len(payload):,} bytes)")


def main():
    parquet_ok = all(os.path.exists(os.path.join(DATA_DIR, f))
                     for f in PARQUET_FILES.values())
    if len(sys.argv) > 1:
        dados = sys.argv[1]
    elif parquet_ok:
        dados = None  # tabelas fato vêm dos parquet em data/
    else:
        dados = find("Dados.xlsx")
    estrutura = sys.argv[2] if len(sys.argv) > 2 else find("Estrutura.xlsx")
    if not estrutura or (not dados and not parquet_ok):
        print("ERRO: coloque os parquet das tabelas fato e Estrutura.xlsx na pasta data/.")
        sys.exit(1)
    print(f"Lendo {dados or 'parquet (data/)'} + {estrutura} ...")

    data = build(dados, estrutura)
    orfaos = build_orfaos(data)
    completar_estrutura(data, orfaos)
    clientes = data.pop("_clientes", None)
    write_all("data.json", data)
    if clientes:
        write_all("clientes.json", clientes)

    write_csv("sem-estrutura.csv",
              [[i["rv"], i["uf"], "-", i.get("cdSv", ""), i.get("cdCv", ""), i["origem"],
                f"{i.get('vl_financeiro', 0.0):.2f}".replace(".", ","),
                f"{i.get('vl_faturado', 0.0):.2f}".replace(".", ","),
                i.get("positivados", 0)] for i in orfaos],
              ["cd_vendedor", "ds_uf", "nome", "cd_vendedor_superior", "cd_gerente", "origem",
               "vl_financeiro", "vl_faturado", "positivados"])
    print(f"OK. generated_at = {data['generated_at']}")
    print(f"comercial: {len(data['comercial'])}  vendas: {len(data['vendas'])}  metas: {len(data['metas'])}")
    print(f"sem estrutura (adicionadas): {len(orfaos)} combinações, "
          f"R$ {sum(i.get('vl_financeiro', 0.0) for i in orfaos):,.2f}")




if __name__ == "__main__":
    main()
